#!/usr/bin/env python3
"""
MMBT Fact-Checking - Experiment Runner
======================================

Unified script to run all experiments:
1. White Noise: Add noise to images and measure impact
2. MuMiN: Integrate and evaluate on MuMiN dataset
3. Complexity: Measure computational complexity vs dataset size

Usage:
    python run_experiments.py --experiment white_noise --noise-levels 0.1,0.2,0.3
    python run_experiments.py --experiment mumin --mumin-size small
    python run_experiments.py --experiment complexity --dataset-sizes 100,500,1000
"""

import argparse
import json
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import torch
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "lib"))

from mmbt.train import Trainer
from mmbt.utils.parseargs import get_args
from mmbt.utils.utils import set_seed
from mmbt.utils.checkpoint_manager import CheckpointManager, get_checkpoint_manager
import configargparse


# ============================================
# Configuration
# ============================================
@dataclass
class ExperimentConfig:
    """Configuration for experiments"""
    experiment_type: str
    output_dir: Path = Path("outputs")
    seed: int = 2024
    device: str = "cuda" if torch.cuda.is_available() else "cpu"
    
    # White noise params
    noise_levels: List[float] = field(default_factory=lambda: [0.0, 0.05, 0.10, 0.20, 0.40])
    
    # MuMiN params
    mumin_size: str = "small"
    mumin_bearer_token: Optional[str] = None
    
    # Complexity params
    dataset_sizes: List[int] = field(default_factory=lambda: [100, 500, 1000, 2000, 5000])
    
    # Dataset
    dataset: str = "Snopes"
    
    # Metrics to track
    metrics: List[str] = field(default_factory=lambda: [
        "ndcg_1", "ndcg_3", "ndcg_5", 
        "hit_1", "hit_3", 
        "training_time", "inference_time"
    ])


# ============================================
# Logging Setup
# ============================================
def setup_logging(output_dir: Path, experiment_name: str, dataset: str = "") -> logging.Logger:
    """Setup logging for experiments"""
    log_dir = output_dir / dataset / "logs" if dataset else output_dir / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"{experiment_name}_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    return logging.getLogger(experiment_name)


# ============================================
# Experiment 1: White Noise
# ============================================
class WhiteNoiseExperiment:
    """
    Experiment to evaluate model robustness to image noise.
    
    Adds Gaussian white noise to images at different levels
    and measures impact on ranking metrics.
    """
    
    def __init__(self, config: ExperimentConfig):
        self.config = config
        self.logger = setup_logging(config.output_dir, "white_noise", config.dataset)
        self.results: Dict[float, Dict] = {}
        
    def add_gaussian_noise(self, image: torch.Tensor, noise_level: float) -> torch.Tensor:
        """
        Add Gaussian white noise to image tensor.
        
        Args:
            image: Input image tensor [C, H, W] or [B, C, H, W]
            noise_level: Standard deviation of noise (0-1)
            
        Returns:
            Noisy image tensor
        """
        noise = torch.randn_like(image) * noise_level
        noisy_image = image + noise
        # Clamp to valid range
        noisy_image = torch.clamp(noisy_image, 0, 1)
        return noisy_image
    
    def create_noisy_transforms(self, noise_level: float):
        """Create transforms that include noise addition"""
        import torchvision.transforms as T
        
        class AddGaussianNoise:
            def __init__(self, std: float):
                self.std = std
                
            def __call__(self, tensor: torch.Tensor) -> torch.Tensor:
                noise = torch.randn_like(tensor) * self.std
                return torch.clamp(tensor + noise, 0, 1)
        
        return T.Compose([
            T.Resize(256),
            T.CenterCrop(224),
            T.ToTensor(),
            T.Normalize(
                mean=[0.46777044, 0.44531429, 0.40661017],
                std=[0.12221994, 0.12145835, 0.14380469],
            ),
            AddGaussianNoise(noise_level),
        ])
    
    def run_single_experiment(self, noise_level: float) -> Dict:
        """Run experiment with specific noise level using the pre-trained checkpoint."""
        import functools
        import torchvision.transforms as T
        from torch.utils.data import DataLoader
        from transformers import BertTokenizer

        from mmbt.data.dataset import TsvDatasetMulti
        from mmbt.data.helpers import collate_fn, get_dataset_annotations_path, get_labels_and_frequencies, get_vocab
        from mmbt.evaluate import Evaluate
        from mmbt.losses.CrossSimilarity import CrossSimilarity
        from mmbt.models import get_model
        from mmbt.utils.utils import load_checkpoint

        self.logger.info(f"Running experiment with noise_level={noise_level}")

        # Use CheckpointManager to locate checkpoint for the chosen dataset
        ckpt_manager = get_checkpoint_manager(str(PROJECT_ROOT / "models" / "checkpoints"))
        dataset_lower = self.config.dataset.lower()
        
        if not ckpt_manager.checkpoint_exists(dataset_lower):
            raise FileNotFoundError(
                f"No checkpoint found for dataset '{dataset_lower}'. "
                f"Run training first: make train CONFIG=configs/{dataset_lower}_config.conf"
            )
        
        checkpoint_dir = ckpt_manager.get_checkpoint_dir(dataset_lower)

        # Restore args saved during training (args.pt contains an argparse.Namespace)
        args = torch.load(checkpoint_dir / "args.pt", weights_only=False)

        # Populate fields that depend on the data (needed by get_model and DataLoader)
        train_path = get_dataset_annotations_path(args, 'train')
        args.labels, args.label_freqs = get_labels_and_frequencies(train_path)
        vocab = get_vocab(args)
        args.vocab = vocab
        args.vocab_sz = vocab.vocab_sz
        args.n_classes = len(args.labels)

        # ---- Noisy transform ------------------------------------------------
        # Noise is applied in pixel-space (after ToTensor, before Normalize)
        # so that clamping to [0,1] is semantically correct.
        class AddGaussianNoise:
            def __init__(self, std: float):
                self.std = std

            def __call__(self, tensor: torch.Tensor) -> torch.Tensor:
                if self.std == 0.0:
                    return tensor
                noise = torch.randn_like(tensor) * self.std
                return torch.clamp(tensor + noise, 0.0, 1.0)

        noisy_transform = T.Compose([
            T.Resize(256),
            T.CenterCrop(224),
            T.ToTensor(),
            AddGaussianNoise(noise_level),
            T.Normalize(
                mean=[0.46777044, 0.44531429, 0.40661017],
                std=[0.12221994, 0.12145835, 0.14380469],
            ),
        ])
        # ---------------------------------------------------------------------

        # Build test DataLoader with the noisy transform
        tokenizer = BertTokenizer.from_pretrained(
            args.bert_model, do_lower_case=True
        ).tokenize

        test_path = get_dataset_annotations_path(args, 'test')
        test_set = TsvDatasetMulti(
            test_path,
            "query",
            tokenizer,
            noisy_transform,
            vocab,
            args,
        )

        collate = functools.partial(collate_fn, args=args)
        test_loader = DataLoader(
            test_set,
            batch_size=args.batch_sz,
            shuffle=False,
            num_workers=args.n_workers,
            collate_fn=collate,
            pin_memory=True,
        )

        # Load model and best checkpoint
        model = get_model(args)
        model.cuda()
        load_checkpoint(model, str(checkpoint_dir / "model_best.pt"))
        model.eval()

        # Run inference
        loss_obj = CrossSimilarity()
        start_time = time.time()
        raw_metrics, _ = Evaluate().model_eval(test_loader, model, args, loss_obj)
        inference_time = time.time() - start_time

        # Map Evaluate keys → report keys
        metrics = {
            "ndcg_1": raw_metrics.get("ndcg_1"),
            "ndcg_3": raw_metrics.get("ndcg_3"),
            "ndcg_5": raw_metrics.get("ndcg_5"),
            "hit_1":  raw_metrics.get("acc_1"),
            "hit_3":  raw_metrics.get("acc_3"),
            "hit_5":  raw_metrics.get("acc_5"),
        }

        self.logger.info(f"noise={noise_level:.2f} | metrics={metrics}")

        return {
            "noise_level": noise_level,
            "metrics": metrics,
            "training_time": 0,
            "inference_time": inference_time,
        }
    
    def run(self) -> None:
        """Run complete white noise experiment"""
        self.logger.info("="*60)
        self.logger.info("Starting White Noise Experiment")
        self.logger.info(f"Noise levels: {self.config.noise_levels}")
        self.logger.info("="*60)
        
        # Run all noise levels (0.0 = baseline)
        for noise_level in self.config.noise_levels:
            self.logger.info(f"\n{'='*40}")
            self.logger.info(f"Noise Level: {noise_level}")
            self.logger.info("="*40)
            
            # Reset seed before each level so all experiments start from
            # the same RNG state — only the noise std differs.
            set_seed(self.config.seed)
            
            results = self.run_single_experiment(noise_level)
            self.results[noise_level] = results
        
        # Save results
        self.save_results()
        
        # Generate report
        self.generate_report()
    
    def save_results(self) -> None:
        """Save experiment results to files"""
        output_dir = self.config.output_dir / self.config.dataset / "white_noise"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Save as JSON
        json_path = output_dir / "results.json"
        with open(json_path, 'w') as f:
            json.dump(self.results, f, indent=2, default=str)
        
        # Save as CSV
        csv_path = output_dir / "results.csv"
        df = pd.DataFrame(self.results).T
        df.to_csv(csv_path)
        
        self.logger.info(f"Results saved to {output_dir}")
    
    def generate_report(self) -> None:
        """Generate markdown report of results"""
        report = ["# White Noise Experiment Results\n"]
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report.append(f"Dataset: {self.config.dataset}\n\n")
        
        report.append("## Results Summary\n")
        report.append("| eta | HIT@3 | HIT@5 | NDCG@3 | NDCG@5 |\n")
        report.append("|-----|-------|-------|--------|--------|\n")
        
        def fmt(val):
            return f"{val:.4f}" if isinstance(val, (int, float)) else str(val)

        for noise_level, results in sorted(self.results.items()):
            metrics = results.get("metrics", {})
            report.append(
                f"| {fmt(noise_level)} | "
                f"{fmt(metrics.get('hit_3', 'N/A'))} | "
                f"{fmt(metrics.get('hit_5', 'N/A'))} | "
                f"{fmt(metrics.get('ndcg_3', 'N/A'))} | "
                f"{fmt(metrics.get('ndcg_5', 'N/A'))} |\n"
            )
        
        output_dir = self.config.output_dir / self.config.dataset / "white_noise"
        report_path = output_dir / "REPORT.md"
        with open(report_path, 'w') as f:
            f.writelines(report)
        
        self.logger.info(f"Report generated: {report_path}")


# ============================================
# Experiment 2: MuMiN Integration
# ============================================
class MuMiNExperiment:
    """
    Experiment to integrate and evaluate on MuMiN dataset.
    
    MuMiN is a large-scale multilingual misinformation dataset
    containing tweets, claims, and fact-checks.
    """
    
    def __init__(self, config: ExperimentConfig):
        self.config = config
        self.logger = setup_logging(config.output_dir, "mumin", config.dataset)
        self.results: Dict = {}
        
    def prepare_dataset(self) -> bool:
        """
        Prepare MuMiN dataset for training.
        
        Downloads and converts MuMiN to format compatible with MMBT.
        """
        self.logger.info("Preparing MuMiN dataset...")
        
        try:
            from mumin import MuminDataset
        except ImportError:
            self.logger.error("MuMiN package not installed. Run: pip install mumin[all]")
            return False
        
        # Get bearer token from config or environment
        bearer_token = self.config.mumin_bearer_token or os.environ.get("TWITTER_BEARER_TOKEN")
        
        if not bearer_token:
            self.logger.error("Twitter Bearer Token not provided. Set TWITTER_BEARER_TOKEN env var.")
            return False
        
        # Initialize dataset
        self.logger.info(f"Loading MuMiN dataset (size={self.config.mumin_size})...")
        dataset = MuminDataset(bearer_token, size=self.config.mumin_size)
        
        # Compile dataset (downloads Twitter data)
        self.logger.info("Compiling dataset (this may take a while)...")
        dataset.compile()
        
        # Convert to MMBT format
        self.logger.info("Converting to MMBT format...")
        self._convert_mumin_to_mmbt(dataset)
        
        return True
    
    def _convert_mumin_to_mmbt(self, dataset) -> None:
        """Convert MuMiN dataset to MMBT TSV format"""
        output_dir = PROJECT_ROOT / "datasets" / "mumin_converted"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Get claims and tweets
        claims_df = dataset.nodes['claim']
        tweets_rel = dataset.rels[('tweet', 'discusses', 'claim')]
        
        # Create train/val/test splits
        train_data = []
        val_data = []
        test_data = []
        
        for idx, claim in claims_df.iterrows():
            # Get associated tweets
            claim_tweets = tweets_rel[tweets_rel['tgt'] == claim['id']]
            
            # Create query-document pairs
            for _, tweet_rel in claim_tweets.iterrows():
                entry = {
                    'QueryID': tweet_rel['src'],
                    'QueryText': '',  # Will be populated from tweet
                    'DocID': claim['id'],
                    'DocText': claim.get('text', ''),
                    'Label': 1 if claim['label'] == 'factual' else 0
                }
                
                # Assign to split based on masks
                if claim.get('small_train_mask', False):
                    train_data.append(entry)
                elif claim.get('small_val_mask', False):
                    val_data.append(entry)
                elif claim.get('small_test_mask', False):
                    test_data.append(entry)
        
        # Save TSV files
        pd.DataFrame(train_data).to_csv(output_dir / "train.tsv", sep='\t', index=False)
        pd.DataFrame(val_data).to_csv(output_dir / "val.tsv", sep='\t', index=False)
        pd.DataFrame(test_data).to_csv(output_dir / "test.tsv", sep='\t', index=False)
        
        self.logger.info(f"Dataset saved to {output_dir}")
        self.logger.info(f"Train: {len(train_data)}, Val: {len(val_data)}, Test: {len(test_data)}")
    
    def run(self) -> None:
        """Run MuMiN experiment"""
        self.logger.info("="*60)
        self.logger.info("Starting MuMiN Integration Experiment")
        self.logger.info(f"Dataset size: {self.config.mumin_size}")
        self.logger.info("="*60)
        
        # Prepare dataset
        if not self.prepare_dataset():
            self.logger.error("Failed to prepare MuMiN dataset")
            return
        
        # Run training
        self.logger.info("Starting training on MuMiN dataset...")
        start_time = time.time()
        
        # TODO: Run actual training
        
        training_time = time.time() - start_time
        self.results["training_time"] = training_time
        
        # Save results
        self.save_results()
    
    def save_results(self) -> None:
        """Save experiment results"""
        output_dir = self.config.output_dir / self.config.dataset / "mumin"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        json_path = output_dir / "results.json"
        with open(json_path, 'w') as f:
            json.dump(self.results, f, indent=2, default=str)
        
        self.logger.info(f"Results saved to {json_path}")


# ============================================
# Experiment 3: Computational Complexity
# ============================================
class ComplexityExperiment:
    """
    Misura la scalabilità del sistema FAISS al crescere della base documentale.

    Strategia:
    - Prende i 7246 documenti Snopes
    - Genera N copie rimuovendo parole a caso (word_removal_ratio=0.25)
    - Codifica i testi con SentenceTransformers (dim=768) → proietta a hidden_sz=128
    - Misura: build index + retrieval k=50 e k=200
    per 7246*10 = 72.460 e 7246*100 = 724.600 documenti
    """
    
    def __init__(self, config: ExperimentConfig):
        self.config = config
        self.logger = setup_logging(config.output_dir, "complexity", config.dataset)
        self.results: Dict = {}
        
    def _load_snopes_documents(self) -> List[str]:
        """Carica i testi dei documenti Snopes da Snopes_DocumentsDataset.tsv"""
        ann_dir = PROJECT_ROOT / "datasets" / "fakenewsdet" / "annotations"
        docs_path = ann_dir / "Snopes_DocumentsDataset.tsv"
        if docs_path.exists():
            df = pd.read_csv(docs_path, sep="\t")
            texts = df["DocText"].dropna().astype(str).tolist()
            self.logger.info(f"Caricati {len(texts)} documenti da {docs_path.name}")
            return texts
        # Fallback: leggi dagli split train/val/test con deduplicazione
        all_docs = []
        for split in ["train.tsv", "val.tsv", "test.tsv"]:
            path = ann_dir / split
            if path.exists():
                df = pd.read_csv(path, sep="\t")
                if "DocText" in df.columns:
                    all_docs.extend(df["DocText"].dropna().astype(str).tolist())
        seen, unique = set(), []
        for t in all_docs:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        self.logger.info(f"Caricati {len(unique)} documenti Snopes unici (fallback)")
        return unique

    def _augment_document(self, text: str, removal_ratio: float = 0.25) -> str:
        """Genera variante rimuovendo parole a caso"""
        words = text.split()
        if len(words) < 4:
            return text
        n_keep = max(2, int(len(words) * (1 - removal_ratio)))
        indices = sorted(np.random.choice(len(words), n_keep, replace=False))
        return " ".join(words[i] for i in indices)

    def _encode_documents(self, docs: List[str], dim: int = 128) -> np.ndarray:
        """
        Codifica i documenti in vettori.
        Usa TF-IDF + SVD per ottenere embedding veloci di dimensione `dim`
        (alternativa leggera a BERT per grandi corpus).
        """
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.decomposition import TruncatedSVD
        
        self.logger.info(f"Encoding {len(docs)} documenti con TF-IDF+SVD (dim={dim})...")
        vectorizer = TfidfVectorizer(max_features=10000, dtype=np.float32)
        X = vectorizer.fit_transform(docs)
        
        n_components = min(dim, X.shape[1] - 1, X.shape[0] - 1)
        svd = TruncatedSVD(n_components=n_components, random_state=42)
        embeddings = svd.fit_transform(X).astype(np.float32)
        
        # Pad o tronca a dim esatto
        if embeddings.shape[1] < dim:
            pad = np.zeros((embeddings.shape[0], dim - embeddings.shape[1]), dtype=np.float32)
            embeddings = np.hstack([embeddings, pad])
        
        # L2 normalize
        norms = np.linalg.norm(embeddings, axis=1, keepdims=True)
        norms = np.where(norms == 0, 1, norms)
        embeddings /= norms
        return embeddings

    def measure_faiss_scaling(
        self, 
        base_docs: List[str],
        expansion_factor: int,
        retrieval_ks: List[int] = [50, 200],
        dim: int = 128,
        word_removal_ratio: float = 0.25,
        n_repetitions: int = 5,
        n_warmup: int = 2,
    ) -> Dict:
        """
        Misura tempo di build FAISS e retrieval per un dato fattore di espansione.
        
        IMPROVEMENTS:
        - Warmup runs to stabilize cache
        - Multiple repetitions with median aggregation
        - Randomized test order to avoid cache bias
        
        Args:
            base_docs: Documenti originali (7246)
            expansion_factor: Quante copie generare (10 o 100)
            retrieval_ks: Valori di k per il retrieval
            dim: Dimensione vettori
            n_repetitions: Numero di ripetizioni per la mediana
            n_warmup: Numero di run di warmup (scartati)
        """
        import faiss
        import random
        
        target_size = len(base_docs) * expansion_factor
        self.logger.info(f"Generazione corpus: {len(base_docs)} base × {expansion_factor} = {target_size} documenti")
        
        # Genera corpus espanso
        expanded_docs = list(base_docs)
        np.random.seed(42)
        for _ in range(expansion_factor - 1):
            augmented = [self._augment_document(d, word_removal_ratio) for d in base_docs]
            expanded_docs.extend(augmented)
        
        # Codifica
        embeddings = self._encode_documents(expanded_docs, dim=dim)
        
        # Query: prime 10 come rappresentative
        queries = embeddings[:10]
        
        results = {
            "n_docs": len(expanded_docs),
            "expansion_factor": expansion_factor,
            "dim": dim,
        }
        
        # === FlatL2 (ricerca esatta) ===
        self.logger.info("Building FlatL2 index...")
        
        # Build time - ripeti 3 volte e prendi mediana
        build_times = []
        for rep in range(3):
            t0 = time.time()
            index_flat = faiss.IndexFlatL2(dim)
            index_flat.add(embeddings)
            build_times.append(time.time() - t0)
        results["flat_index_build_time"] = np.median(build_times)
        self.logger.info(f"  Build time (median of 3): {results['flat_index_build_time']:.4f}s ± {np.std(build_times):.4f}s")
        
        # Ricostruisci indice finale
        index_flat = faiss.IndexFlatL2(dim)
        index_flat.add(embeddings)
        
        # Search time - con warmup e ripetizioni
        for k in retrieval_ks:
            # Warmup
            for _ in range(n_warmup):
                index_flat.search(queries, k)
            
            # Misurazioni effettive
            search_times = []
            for _ in range(n_repetitions):
                t0 = time.time()
                _, _ = index_flat.search(queries, k)
                search_times.append(time.time() - t0)
            
            median_time = np.median(search_times)
            std_time = np.std(search_times)
            results[f"flat_search_k{k}_time"] = median_time
            results[f"flat_search_k{k}_per_query_ms"] = (median_time / len(queries)) * 1000
            results[f"flat_search_k{k}_std"] = std_time
            
            self.logger.info(
                f"  FlatL2 k={k}: {median_time:.4f}s ± {std_time:.4f}s "
                f"({results[f'flat_search_k{k}_per_query_ms']:.2f} ms/query)"
            )
        
        # === IVFFlat (ricerca approssimata) ===
        nlist = min(100, max(10, len(expanded_docs) // 1000))
        self.logger.info(f"Building IVFFlat index (nlist={nlist})...")
        
        # Build time - ripeti 3 volte e prendi mediana
        build_times = []
        for rep in range(3):
            t0 = time.time()
            quantizer = faiss.IndexFlatL2(dim)
            index_ivf = faiss.IndexIVFFlat(quantizer, dim, nlist)
            index_ivf.train(embeddings)
            index_ivf.add(embeddings)
            build_times.append(time.time() - t0)
        results["ivf_index_build_time"] = np.median(build_times)
        self.logger.info(f"  IVF Build time (median of 3): {results['ivf_index_build_time']:.4f}s ± {np.std(build_times):.4f}s")
        
        # Ricostruisci indice finale
        quantizer = faiss.IndexFlatL2(dim)
        index_ivf = faiss.IndexIVFFlat(quantizer, dim, nlist)
        index_ivf.train(embeddings)
        index_ivf.add(embeddings)
        
        # Search time - con warmup e ripetizioni, ordine randomizzato
        test_configs = [(k, rep) for k in retrieval_ks for rep in range(n_warmup + n_repetitions)]
        random.shuffle(test_configs)
        
        search_results = {k: [] for k in retrieval_ks}
        for k, rep in test_configs:
            t0 = time.time()
            _, _ = index_ivf.search(queries, k)
            elapsed = time.time() - t0
            
            # Scarta warmup, raccogli solo le misurazioni reali
            if rep >= n_warmup:
                search_results[k].append(elapsed)
        
        for k in retrieval_ks:
            median_time = np.median(search_results[k])
            std_time = np.std(search_results[k])
            results[f"ivf_search_k{k}_time"] = median_time
            results[f"ivf_search_k{k}_per_query_ms"] = (median_time / len(queries)) * 1000
            results[f"ivf_search_k{k}_std"] = std_time
            
            self.logger.info(
                f"  IVFFlat k={k}: {median_time:.4f}s ± {std_time:.4f}s "
                f"({results[f'ivf_search_k{k}_per_query_ms']:.2f} ms/query)"
            )
        
        return results

    def run(self) -> None:
        """Esegue l'esperimento di scalabilità"""
        self.logger.info("=" * 60)
        self.logger.info("Starting Scalability Experiment")
        self.logger.info("=" * 60)
        
        # Carica documenti base
        base_docs = self._load_snopes_documents()
        dim = 128  # hidden_sz del modello
        
        expansion_factors = [10, 100]
        retrieval_ks = [50, 200]
        
        for factor in expansion_factors:
            self.logger.info(f"\n{'='*40}")
            self.logger.info(f"Expansion factor: {factor}x ({len(base_docs)*factor:,} documenti)")
            self.logger.info("=" * 40)
            result = self.measure_faiss_scaling(
                base_docs=base_docs,
                expansion_factor=factor,
                retrieval_ks=retrieval_ks,
                dim=dim,
            )
            self.results[factor] = result
        
        self.save_results()
        self.generate_report()
        self.generate_excel()

    def save_results(self) -> None:
        """Salva i risultati"""
        output_dir = self.config.output_dir / self.config.dataset / "complexity"
        output_dir.mkdir(parents=True, exist_ok=True)
        json_path = output_dir / "scalability.json"
        with open(json_path, 'w') as f:
            json.dump(self.results, f, indent=2, default=str)
        self.logger.info(f"Results saved to {json_path}")

    def generate_report(self) -> None:
        """Genera report con tabella risultati"""
        report = ["# Scalability Analysis - FAISS\n\n"]
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        report.append("Base dataset: Snopes (1,703 documenti)\n")
        report.append("Varianti generate rimuovendo 25% parole a caso\n\n")
        report.append("**Measurement protocol**:\n")
        report.append("- Build times: median of 3 repetitions\n")
        report.append("- Search times: median of 5 repetitions with 2 warmup runs\n")
        report.append("- Query set: 10 representative queries\n\n")

        report.append("## FlatL2 (Ricerca Esatta)\n\n")
        report.append("| N Documenti | Build (s) | k=50 (ms/q) | k=50 std | k=200 (ms/q) | k=200 std |\n")
        report.append("|------------|-----------|-------------|---------|--------------|----------|\n")
        for factor, data in sorted(self.results.items()):
            report.append(
                f"| {data['n_docs']:,} | "
                f"{data.get('flat_index_build_time', 0):.4f} | "
                f"{data.get('flat_search_k50_per_query_ms', 0):.3f} | "
                f"±{data.get('flat_search_k50_std', 0)*1000/10:.3f} | "  # std per query in ms
                f"{data.get('flat_search_k200_per_query_ms', 0):.3f} | "
                f"±{data.get('flat_search_k200_std', 0)*1000/10:.3f} |\n"
            )

        report.append("\n## IVFFlat (Ricerca Approssimata)\n\n")
        report.append("| N Documenti | Build (s) | k=50 (ms/q) | k=50 std | k=200 (ms/q) | k=200 std |\n")
        report.append("|------------|-----------|-------------|---------|--------------|----------|\n")
        for factor, data in sorted(self.results.items()):
            report.append(
                f"| {data['n_docs']:,} | "
                f"{data.get('ivf_index_build_time', 0):.4f} | "
                f"{data.get('ivf_search_k50_per_query_ms', 0):.3f} | "
                f"±{data.get('ivf_search_k50_std', 0)*1000/10:.3f} | "  # std per query in ms
                f"{data.get('ivf_search_k200_per_query_ms', 0):.3f} | "
                f"±{data.get('ivf_search_k200_std', 0)*1000/10:.3f} |\n"
            )
        
        # Aggiungi sezione di analisi comparativa
        report.append("\n## Performance Comparison\n\n")
        for factor, data in sorted(self.results.items()):
            report.append(f"### Corpus size: {data['n_docs']:,} documents (expansion {factor}x)\n\n")
            
            # Speedup IVF vs Flat
            flat_k50 = data.get('flat_search_k50_per_query_ms', 1)
            ivf_k50 = data.get('ivf_search_k50_per_query_ms', 1)
            flat_k200 = data.get('flat_search_k200_per_query_ms', 1)
            ivf_k200 = data.get('ivf_search_k200_per_query_ms', 1)
            
            speedup_k50 = flat_k50 / ivf_k50 if ivf_k50 > 0 else 0
            speedup_k200 = flat_k200 / ivf_k200 if ivf_k200 > 0 else 0
            
            report.append(f"- **IVF Speedup (k=50)**: {speedup_k50:.1f}× faster than FlatL2\n")
            report.append(f"- **IVF Speedup (k=200)**: {speedup_k200:.1f}× faster than FlatL2\n")
            
            # Build time overhead
            flat_build = data.get('flat_index_build_time', 0)
            ivf_build = data.get('ivf_index_build_time', 0)
            build_overhead = ivf_build / flat_build if flat_build > 0 else 0
            
            report.append(f"- **Build Time Overhead**: IVF takes {build_overhead:.1f}× longer to build\n")
            
            # Break-even point (quante query servono per ammortizzare il build time)
            if ivf_k50 > 0 and flat_k50 > ivf_k50:
                time_saved_per_query = (flat_k50 - ivf_k50) / 1000  # in secondi
                queries_to_breakeven = (ivf_build - flat_build) / time_saved_per_query
                report.append(f"- **Break-even Point**: ~{int(queries_to_breakeven)} queries to amortize build cost\n")
            
            report.append("\n")
        
        output_dir = self.config.output_dir / self.config.dataset / "complexity"
        report_path = output_dir / "REPORT.md"
        with open(report_path, 'w') as f:
            f.writelines(report)
        self.logger.info(f"Report: {report_path}")

    def generate_excel(self) -> None:
        """Genera file Excel con formato compatibile per il paper"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Foglio1"
        
        # Stile intestazioni
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        # === Tabella FlatL2 ===
        row = 2
        headers_flat = ["Expansion Factor", "n_docs", "dim", 
                       "Flat Build Time (s)", "Flat Search k=50 (s)", "Flat Search k=200 (s)"]
        
        for col, header in enumerate(headers_flat, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        row += 1
        for factor, data in sorted(self.results.items()):
            ws.cell(row=row, column=1, value=int(factor))
            ws.cell(row=row, column=2, value=data['n_docs'])
            ws.cell(row=row, column=3, value=data['dim'])
            ws.cell(row=row, column=4, value=round(data.get('flat_index_build_time', 0), 6))
            ws.cell(row=row, column=5, value=round(data.get('flat_search_k50_time', 0), 6))
            ws.cell(row=row, column=6, value=round(data.get('flat_search_k200_time', 0), 6))
            row += 1
        
        # Riga vuota
        row += 1
        
        # === Tabella IVF ===
        headers_ivf = ["Expansion Factor", "n_docs", "dim",
                      "IVF Build Time (s)", "IVF Search k=50 (s)", "IVF Search k=200 (s)"]
        
        for col, header in enumerate(headers_ivf, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        row += 1
        for factor, data in sorted(self.results.items()):
            ws.cell(row=row, column=1, value=int(factor))
            ws.cell(row=row, column=2, value=data['n_docs'])
            ws.cell(row=row, column=3, value=data['dim'])
            ws.cell(row=row, column=4, value=round(data.get('ivf_index_build_time', 0), 6))
            ws.cell(row=row, column=5, value=round(data.get('ivf_search_k50_time', 0), 6))
            ws.cell(row=row, column=6, value=round(data.get('ivf_search_k200_time', 0), 6))
            row += 1
        
        # Imposta larghezza colonne
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # Salva file
        output_dir = self.config.output_dir / self.config.dataset / "complexity"
        excel_path = output_dir / f"{self.config.dataset.upper()}-Scalabilita_FAISS_Complexity_MMBT.xlsx"
        wb.save(excel_path)
        self.logger.info(f"Excel file: {excel_path}")



# ============================================
# Main Entry Point
# ============================================
def parse_arguments() -> argparse.Namespace:
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description="MMBT Fact-Checking Experiment Runner",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    # Run white noise experiment
    python run_experiments.py --experiment white_noise --noise-levels 0.1,0.2,0.3,0.5

    # Run MuMiN experiment  
    python run_experiments.py --experiment mumin --mumin-size small

    # Run complexity analysis
    python run_experiments.py --experiment complexity --dataset-sizes 100,500,1000,2000

    # Run evaluation only
    python run_experiments.py --experiment evaluate --dataset Snopes
        """
    )
    
    parser.add_argument(
        "--experiment",
        type=str,
        required=True,
        choices=["white_noise", "mumin", "complexity", "evaluate", "all"],
        help="Type of experiment to run"
    )
    
    parser.add_argument(
        "--output-dir",
        type=str,
        default="outputs",
        help="Directory to save outputs"
    )
    
    parser.add_argument(
        "--seed",
        type=int,
        default=2024,
        help="Random seed for reproducibility"
    )
    
    parser.add_argument(
        "--dataset",
        type=str,
        default="Snopes",
        choices=["Snopes", "Politifact", "MuMiN", "M3Check", "M3CheckNoEng", "snopes", "politifact", "mumin", "m3check", "m3check_multilingual", "m3check_no_eng"],
        help="Dataset to use"
    )
    
    # White noise arguments
    parser.add_argument(
        "--noise-levels",
        type=str,
        default="0.0,0.05,0.10,0.20,0.40",
        help="Comma-separated noise levels for white noise experiment"
    )
    
    # MuMiN arguments
    parser.add_argument(
        "--mumin-size",
        type=str,
        default="small",
        choices=["small", "medium", "large"],
        help="MuMiN dataset size"
    )
    
    parser.add_argument(
        "--mumin-bearer-token",
        type=str,
        default=None,
        help="Twitter Bearer Token for MuMiN (or set TWITTER_BEARER_TOKEN env)"
    )
    
    # Complexity arguments
    parser.add_argument(
        "--dataset-sizes",
        type=str,
        default="100,500,1000,2000,5000",
        help="Comma-separated dataset sizes for complexity analysis"
    )
    
    # Config file override
    parser.add_argument(
        "--config",
        type=str,
        default=None,
        help="YAML config file to override arguments"
    )
    
    return parser.parse_args()


def main():
    """Main entry point"""
    args = parse_arguments()
    
    # Set seed for reproducibility
    set_seed(args.seed)
    
    # Create config
    config = ExperimentConfig(
        experiment_type=args.experiment,
        output_dir=Path(args.output_dir),
        seed=args.seed,
        dataset=args.dataset,
        noise_levels=[float(x) for x in args.noise_levels.split(",")],
        mumin_size=args.mumin_size,
        mumin_bearer_token=args.mumin_bearer_token,
        dataset_sizes=[int(x) for x in args.dataset_sizes.split(",")]
    )
    
    # Override with config file if provided
    if args.config:
        with open(args.config, 'r') as f:
            yaml_config = yaml.safe_load(f)
            for key, value in yaml_config.items():
                if hasattr(config, key):
                    setattr(config, key, value)
    
    # Run appropriate experiment
    print(f"\n{'='*60}")
    print(f"MMBT Fact-Checking - Experiment: {args.experiment.upper()}")
    print(f"{'='*60}\n")
    
    if args.experiment == "white_noise":
        experiment = WhiteNoiseExperiment(config)
        experiment.run()
        
    elif args.experiment == "mumin":
        experiment = MuMiNExperiment(config)
        experiment.run()
        
    elif args.experiment == "complexity":
        experiment = ComplexityExperiment(config)
        experiment.run()
        
    elif args.experiment == "evaluate":
        # Run evaluation using CheckpointManager
        print(f"Running evaluation for {config.dataset}...")
        ckpt_manager = get_checkpoint_manager(str(PROJECT_ROOT / "models" / "checkpoints"))
        
        # Check checkpoint exists
        dataset_lower = config.dataset.lower()
        if not ckpt_manager.checkpoint_exists(dataset_lower):
            print(f"❌ No checkpoint found for {dataset_lower}")
            print(f"   Train first with: make train CONFIG=configs/{dataset_lower}_config.conf")
            sys.exit(1)
        
        print(f"✅ Using checkpoint: {ckpt_manager.get_checkpoint_dir(dataset_lower)}")
        
        # Run baseline evaluation (noise_level=0)
        noise_config = ExperimentConfig(
            experiment_type="white_noise",
            output_dir=config.output_dir,
            seed=config.seed,
            dataset=config.dataset,
            noise_levels=[0.0]
        )
        eval_exp = WhiteNoiseExperiment(noise_config)
        eval_exp.run()
        
    elif args.experiment == "all":
        print("Running all experiments...")
        
        # White noise
        noise_exp = WhiteNoiseExperiment(config)
        noise_exp.run()
        
        # Complexity
        complexity_exp = ComplexityExperiment(config)
        complexity_exp.run()
        
        # MuMiN (optional, requires credentials)
        if config.mumin_bearer_token:
            mumin_exp = MuMiNExperiment(config)
            mumin_exp.run()
    
    print(f"\n{'='*60}")
    print("Experiment completed!")
    print(f"Results saved to: {config.output_dir}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
